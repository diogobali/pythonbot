import selenium
from selenium import webdriver
import time
import cv2
import pytesseract
from PIL import Image
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By


cont = 0
contAux = 0
contCel = 2
cpfTemp = '0'
nomeTemp = 'abc'
orgaoTemp = 'abc'
contbrutaCel = 2
erro = 'abc'
login = 'abc'
senha = 'abc'
captcha = 'abc'



book = Workbook()
sheet = book.active


sheet['A1'] = 'CPF'
sheet['B1'] = 'NOME'
sheet['D1'] = 'ORGAO'
sheet['M1'] = 'MARGEM BRUTA CONSIGNAVEL'
sheet['O1'] = 'MARGEM BRUTA CARTAO'
sheet['U1'] = 'MARGEM CONSIGNAVEL'
sheet['W1'] = 'MARGEM CARTAO'

nomebasebruta = input("Digite o nome da base bruta (incluindo extensão .xlsx): ")
nome_xlsx = 'resultado'+nomebasebruta
basebruta = load_workbook(nomebasebruta)
sheetbruta = basebruta['Planilha1']
sheet.column_dimensions['C'].hidden= True
sheet.column_dimensions['E'].hidden= True
sheet.column_dimensions['F'].hidden= True
sheet.column_dimensions['G'].hidden= True
sheet.column_dimensions['H'].hidden= True
sheet.column_dimensions['J'].hidden= True
sheet.column_dimensions['K'].hidden= True
sheet.column_dimensions['L'].hidden= True
sheet.column_dimensions['N'].hidden= True
sheet.column_dimensions['P'].hidden= True
sheet.column_dimensions['R'].hidden= True
sheet.column_dimensions['S'].hidden= True
sheet.column_dimensions['T'].hidden= True
sheet.column_dimensions['V'].hidden= True

contAux = int(input("Digite a quantidade de consultas que deseja: "))
login = input("Digite o login do portal: ")
senha = input("Digite a senha do portal: ")




driver = webdriver.Chrome()
driver.get('https://www.portaldoconsignado.com.br/home?0')
time.sleep(4)
driver.find_element_by_xpath('//*[@id="guias"]/div[2]/span/span').click()
time.sleep(2)
driver.find_element_by_id('username').click()
username = driver.find_element_by_id('username').send_keys(login)
password = driver.find_element_by_id('password')
time.sleep(2)
password.send_keys(senha)
# ATÉ AQUI COLOCAMOS USUARIO E SENHA #
captcha = input("Digite o captcha do portal: ")
captcha = driver.find_element_by_id('captcha').send_keys(captcha)
driver.find_element_by_xpath('//*[@value="Acessar"]').click()
time.sleep(10)
#driver.find_element_by_xpath('//*[@id="divlistaPerfil"]/fieldset/span/label/span').click()
#time.sleep(2)
#driver.find_element_by_xpath('//*[@id="id16"]').click()
#time.sleep(4)
while (cont < contAux):
    cpfbrutoTemp = sheetbruta['A'+str(contbrutaCel)].value
    driver.get('https://www.portaldoconsignado.com.br/consignatario/pesquisarMargem?7')
    element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'main')))
    driver.find_element_by_id('cpfServidor').clear()
    time.sleep(2)
    element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'main')))
    element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'cpfServidor')))
    driver.find_element_by_id('cpfServidor').send_keys(cpfbrutoTemp)
    element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@value="Pesquisar"]')))
    driver.find_element_by_xpath('//*[@value="Pesquisar"]').click()
    time.sleep(6)
    try:
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'divResultadoServidor')))
        cpfTemp = driver.find_element_by_id('divResultadoServidor')
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'painelMargensBrutas')))
        margemTemp = driver.find_element_by_id('painelMargensBrutas')
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'painelMargensDisponiveis')))
        margemdispTemp = driver.find_element_by_id('painelMargensDisponiveis')
        sheet['A'+str(contCel)] = cpfTemp.text
        sheet['I'+str(contCel)] = margemTemp.text
        sheet['Q'+str(contCel)] = margemdispTemp.text
        print("Consultando "+str(contCel-1)+"| CPF "+str(cpfbrutoTemp))
            # nomeTemp = driver.find_element_by_xpath('//*[@id="id7a"]')
            # orgaoTemp = driver.find_element_by_xpath('//*[@id="id7c"]')
            # sheet['B'+str(contCel)] = nomeTemp
            # sheet['C'+str(contCel)] = orgaoTemp
    except:
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'divEtapaError2')))
        erro = driver.find_element_by_id('divEtapaError2')
        sheet['I'+str(contCel)] = erro.text 
        sheet['A'+str(contCel)] = cpfbrutoTemp
        print("Consultando "+str(contCel-1)+"| CPF "+str(cpfbrutoTemp))
    cont = cont + 1
    contCel = contCel + 1
    contbrutaCel = contbrutaCel + 1
book.save(nome_xlsx)
    


